import os
import io
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import db, Room, Receptionist, Checklist
from rooms_data import get_buildings
from sqlalchemy import func, cast, Date

app = Flask(__name__)

# Database config
database_url = os.environ.get('DATABASE_URL', 'sqlite:///camp_checklist.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True
}
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key')

# Delete password (configurable via env var)
DELETE_PASSWORD = os.environ.get('DELETE_PASSWORD', 'admin2026')

db.init_app(app)

with app.app_context():
    db.create_all()


# ── Helpers ────────────────────────────────────────────────────────────────

def get_dashboard_range():
    """Resolve dashboard date filtering."""
    selected_date = (request.args.get('selected_date') or '').strip()
    days = request.args.get('days', 7, type=int)

    if selected_date:
        try:
            start_date = datetime.strptime(selected_date, '%Y-%m-%d')
            end_date = start_date + timedelta(days=1)
            return {
                'start_date': start_date,
                'end_date': end_date,
                'selected_date': selected_date,
                'days': None,
                'label': selected_date
            }
        except ValueError:
            pass

    start_date = datetime.utcnow() - timedelta(days=days)
    return {
        'start_date': start_date,
        'end_date': None,
        'selected_date': '',
        'days': days,
        'label': f'Últimos {days} día(s)'
    }


# ── Pages ──────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    receptionists = Receptionist.query.filter_by(active=True).order_by(Receptionist.name).all()
    buildings = get_buildings()
    return render_template('index.html', receptionists=receptionists, buildings=buildings)


@app.route('/checklist/<building>')
def checklist_page(building):
    receptionist_id = request.args.get('receptionist_id', type=int)
    if not receptionist_id:
        return redirect(url_for('index'))
    receptionist = Receptionist.query.get_or_404(receptionist_id)
    rooms = Room.query.filter_by(building=building).order_by(Room.code).all()
    return render_template(
        'checklist.html',
        rooms=rooms,
        building=building,
        receptionist=receptionist
    )


@app.route('/dashboard')
def dashboard_page():
    return render_template('dashboard.html')


@app.route('/receptionists')
def receptionists_page():
    receptionists = Receptionist.query.order_by(Receptionist.name).all()
    return render_template('receptionists.html', receptionists=receptionists)


@app.route('/history')
def history_page():
    return render_template('history.html')


# ── API: Receptionists ────────────────────────────────────────────────────

@app.route('/api/receptionists', methods=['GET'])
def api_get_receptionists():
    recs = Receptionist.query.filter_by(active=True).order_by(Receptionist.name).all()
    return jsonify([r.to_dict() for r in recs])


@app.route('/api/receptionists', methods=['POST'])
def api_create_receptionist():
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'El nombre es requerido'}), 400
    existing = Receptionist.query.filter_by(name=name).first()
    if existing:
        if not existing.active:
            existing.active = True
            db.session.commit()
            return jsonify(existing.to_dict()), 200
        return jsonify({'error': 'Ya existe un recepcionista con ese nombre'}), 400
    rec = Receptionist(name=name)
    db.session.add(rec)
    db.session.commit()
    return jsonify(rec.to_dict()), 201


@app.route('/api/receptionists/<int:rec_id>', methods=['PUT'])
def api_update_receptionist(rec_id):
    rec = Receptionist.query.get_or_404(rec_id)
    data = request.json
    if 'name' in data:
        rec.name = data['name'].strip()
    if 'active' in data:
        rec.active = data['active']
    db.session.commit()
    return jsonify(rec.to_dict())


@app.route('/api/receptionists/<int:rec_id>', methods=['DELETE'])
def api_delete_receptionist(rec_id):
    rec = Receptionist.query.get_or_404(rec_id)
    rec.active = False
    db.session.commit()
    return jsonify({'ok': True})


# ── API: Checklist ─────────────────────────────────────────────────────────

@app.route('/api/checklist', methods=['POST'])
def api_create_checklist():
    data = request.json
    room_id = data.get('room_id')
    receptionist_id = data.get('receptionist_id')

    if not room_id or not receptionist_id:
        return jsonify({'error': 'room_id y receptionist_id son requeridos'}), 400

    checklist = Checklist(
        room_id=room_id,
        receptionist_id=receptionist_id,
        disponibilidad_cupos=data.get('disponibilidad_cupos', ''),
        limpieza_general=data.get('limpieza_general', ''),
        limpieza_banos=data.get('limpieza_banos', ''),
        insumos_basicos=data.get('insumos_basicos', ''),
        iluminacion=data.get('iluminacion', ''),
        agua=data.get('agua', ''),
        ventanas=data.get('ventanas', ''),
        cortinas=data.get('cortinas', ''),
        estufas=data.get('estufas', ''),
        mobiliario=data.get('mobiliario', ''),
        chapas=data.get('chapas', ''),
        cambio_sabanas=data.get('cambio_sabanas', ''),
        casilleros=data.get('casilleros', ''),
        observaciones=data.get('observaciones', '')
    )
    db.session.add(checklist)
    db.session.commit()
    return jsonify(checklist.to_dict()), 201


@app.route('/api/checklist/batch', methods=['POST'])
def api_create_checklist_batch():
    """Save multiple room checklists at once."""
    data = request.json
    items = data.get('items', [])
    receptionist_id = data.get('receptionist_id')

    if not items or not receptionist_id:
        return jsonify({'error': 'items y receptionist_id son requeridos'}), 400

    saved = []
    for item in items:
        checklist = Checklist(
            room_id=item['room_id'],
            receptionist_id=receptionist_id,
            disponibilidad_cupos=item.get('disponibilidad_cupos', ''),
            limpieza_general=item.get('limpieza_general', ''),
            limpieza_banos=item.get('limpieza_banos', ''),
            insumos_basicos=item.get('insumos_basicos', ''),
            iluminacion=item.get('iluminacion', ''),
            agua=item.get('agua', ''),
            ventanas=item.get('ventanas', ''),
            cortinas=item.get('cortinas', ''),
            estufas=item.get('estufas', ''),
            mobiliario=item.get('mobiliario', ''),
            chapas=item.get('chapas', ''),
            cambio_sabanas=item.get('cambio_sabanas', ''),
            casilleros=item.get('casilleros', ''),
            observaciones=item.get('observaciones', '')
        )
        db.session.add(checklist)
        saved.append(checklist)

    db.session.commit()
    return jsonify({'saved': len(saved)}), 201


@app.route('/api/checklist/<int:checklist_id>', methods=['DELETE'])
def api_delete_checklist(checklist_id):
    """Delete a checklist record with password verification."""
    data = request.json or {}
    password = data.get('password', '')
    if password != DELETE_PASSWORD:
        return jsonify({'error': 'Clave incorrecta'}), 403
    checklist = Checklist.query.get_or_404(checklist_id)
    db.session.delete(checklist)
    db.session.commit()
    return jsonify({'ok': True})


# ── API: Dashboard ─────────────────────────────────────────────────────────

@app.route('/api/dashboard/stats')
def api_dashboard_stats():
    """Dashboard statistics."""
    date_filter = get_dashboard_range()
    start_date = date_filter['start_date']
    end_date = date_filter['end_date']

    is_sqlite = 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']
    if is_sqlite:
        date_expr = func.date(Checklist.created_at)
    else:
        date_expr = cast(Checklist.created_at, Date)

    base_filters = [Checklist.created_at >= start_date]
    if end_date is not None:
        base_filters.append(Checklist.created_at < end_date)

    daily = db.session.query(
        date_expr.label('date'),
        func.count(Checklist.id).label('count')
    ).filter(
        *base_filters
    ).group_by(
        date_expr
    ).order_by(
        date_expr
    ).all()

    by_receptionist = db.session.query(
        Receptionist.name,
        func.count(Checklist.id).label('count')
    ).join(
        Checklist, Checklist.receptionist_id == Receptionist.id
    ).filter(
        *base_filters
    ).group_by(
        Receptionist.name
    ).order_by(
        func.count(Checklist.id).desc()
    ).all()

    issue_fields = [
        'limpieza_general', 'limpieza_banos', 'insumos_basicos',
        'iluminacion', 'agua', 'ventanas', 'cortinas',
        'estufas', 'mobiliario', 'chapas', 'cambio_sabanas'
    ]

    field_labels = {
        'limpieza_general': 'Limpieza General',
        'limpieza_banos': 'Limpieza Baños',
        'insumos_basicos': 'Insumos Básicos',
        'iluminacion': 'Iluminación',
        'agua': 'Agua',
        'ventanas': 'Ventanas',
        'cortinas': 'Cortinas',
        'estufas': 'Estufas',
        'mobiliario': 'Mobiliario',
        'chapas': 'Chapas',
        'cambio_sabanas': 'Cambio Sábanas'
    }

    issues_data = {}
    for field in issue_fields:
        count = Checklist.query.filter(
            *base_filters,
            getattr(Checklist, field) == 'x'
        ).count()
        issues_data[field_labels[field]] = count

    total_checklists = Checklist.query.filter(*base_filters).count()

    total_rooms = Room.query.count()
    rooms_checked = db.session.query(
        func.count(func.distinct(Checklist.room_id))
    ).filter(
        *base_filters
    ).scalar()

    by_building = db.session.query(
        Room.building,
        func.count(Checklist.id).label('count')
    ).join(
        Checklist, Checklist.room_id == Room.id
    ).filter(
        *base_filters
    ).group_by(
        Room.building
    ).order_by(
        Room.building
    ).all()

    return jsonify({
        'daily': [{'date': str(d.date), 'count': d.count} for d in daily],
        'by_receptionist': [{'name': r.name, 'count': r.count} for r in by_receptionist],
        'issues': issues_data,
        'total_checklists': total_checklists,
        'total_rooms': total_rooms,
        'rooms_checked': rooms_checked or 0,
        'by_building': [{'building': b.building, 'count': b.count} for b in by_building],
        'selected_date': date_filter['selected_date'],
        'days': date_filter['days'],
        'filter_label': date_filter['label']
    })


# ── API: History ───────────────────────────────────────────────────────────

@app.route('/api/history')
def api_history():
    """Get checklist history with filters."""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    building = request.args.get('building', '')
    receptionist_id = request.args.get('receptionist_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = Checklist.query.join(Room).join(Receptionist)

    if building:
        query = query.filter(Room.building == building)
    if receptionist_id:
        query = query.filter(Checklist.receptionist_id == receptionist_id)
    if date_from:
        query = query.filter(Checklist.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        query = query.filter(Checklist.created_at < datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))

    total = query.count()
    checklists = query.order_by(Checklist.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    return jsonify({
        'items': [c.to_dict() for c in checklists],
        'total': total,
        'page': page,
        'pages': (total + per_page - 1) // per_page
    })


@app.route('/api/history/export')
def api_history_export():
    """Export history to Excel (.xlsx)."""
    building = request.args.get('building', '')
    receptionist_id = request.args.get('receptionist_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = Checklist.query.join(Room).join(Receptionist)

    if building:
        query = query.filter(Room.building == building)
    if receptionist_id:
        query = query.filter(Checklist.receptionist_id == receptionist_id)
    if date_from:
        query = query.filter(Checklist.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        query = query.filter(Checklist.created_at < datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))

    checklists = query.order_by(Checklist.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Chequeos'

    headers = [
        'Fecha', 'Hora', 'Habitación', 'Sector', 'Recepcionista',
        'Cupos', 'Limpieza Gral', 'Baños', 'Insumos', 'Iluminación',
        'Agua', 'Ventanas', 'Cortinas', 'Estufas', 'Mobiliario',
        'Chapas', 'Cambio Sábanas', 'Casilleros', 'Observaciones'
    ]

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='B42318', end_color='B42318', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ok_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    x_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')

    for row_idx, c in enumerate(checklists, 2):
        values = [
            c.created_at.strftime('%Y-%m-%d'),
            c.created_at.strftime('%H:%M:%S'),
            c.room.code,
            c.room.building,
            c.receptionist.name,
            c.disponibilidad_cupos.upper() if c.disponibilidad_cupos else '',
            c.limpieza_general.upper() if c.limpieza_general else '',
            c.limpieza_banos.upper() if c.limpieza_banos else '',
            c.insumos_basicos.upper() if c.insumos_basicos else '',
            c.iluminacion.upper() if c.iluminacion else '',
            c.agua.upper() if c.agua else '',
            c.ventanas.upper() if c.ventanas else '',
            c.cortinas.upper() if c.cortinas else '',
            c.estufas.upper() if c.estufas else '',
            c.mobiliario.upper() if c.mobiliario else '',
            c.chapas.upper() if c.chapas else '',
            c.cambio_sabanas.upper() if c.cambio_sabanas else '',
            c.casilleros.upper() if c.casilleros else '',
            c.observaciones or ''
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center_align
            if 7 <= col_idx <= 17:
                if val == 'OK':
                    cell.fill = ok_fill
                elif val == 'X':
                    cell.fill = x_fill

    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, min(len(checklists) + 2, 100)):
            val = ws.cell(row=row, column=col).value
            if val and len(str(val)) > max_len:
                max_len = len(str(val))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 30)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(checklists) + 1}'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'chequeos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
